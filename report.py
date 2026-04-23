# report.py - Monthly Attendance Report with Excel export
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os, sys, calendar, threading
from datetime import date, datetime, timedelta
from PIL import Image, ImageTk, ImageFilter, ImageEnhance
from database import Database
from gradient import GradientFrame, GradientButton
import theme as TM
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_base_dir():
    if getattr(sys, 'frozen', False):
        # PyInstaller stores files in _MEIPASS when running as .exe
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))
def _set_app_icon(win):
    """Set favicon.png as window icon."""
    base = get_base_dir()
    for name in ["favicon.png", "favicon.ico"]:
        path = os.path.join(base, name)
        if os.path.exists(path):
            try:
                img = tk.PhotoImage(file=path)
                win.iconphoto(False, img)
                win._icon_ref = img  # prevent GC
            except: pass
            break



def mg(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


# ── Excel export ──────────────────────────────────────────────────────────────

def generate_excel_report(output_path, semester, month, year,
                           room, teacher_name, subject,
                           students, attendance_data):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_name[month][:3]} {year} Sem{semester}"

    HEADER_BG = "4472C4"
    SUN_BG    = "BDD7EE"
    ALT_BG    = "F2F2F2"
    BORDER_C  = "BFBFBF"
    thin = Side(style='thin', color=BORDER_C)
    bd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def mf(sz=9, bold=False, color='000000'):
        return Font(name='Arial', size=sz, bold=bold, color=color)
    def fl(c):
        return PatternFill('solid', start_color=c, fgColor=c)
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left',   vertical='center')

    num_days  = calendar.monthrange(year, month)[1]
    day_abbrs = [calendar.day_abbr[date(year, month, d).weekday()][:3]
                 for d in range(1, num_days+1)]
    sundays   = {d for d in range(1, num_days+1)
                 if date(year, month, d).weekday() == 6}

    NO_COL = 1; NAME_COL = 2; DS = 3
    last_col = DS + num_days - 1

    # Title
    mg(ws, 1, 1, 1, last_col)
    c = ws.cell(1, 1, "Monthly Attendance Report")
    c.font = mf(16, True, 'FFFFFF'); c.fill = fl(HEADER_BG); c.alignment = ctr
    ws.row_dimensions[1].height = 32

    # Meta rows 2-4
    rows_meta = [
        [("Room:", room),
         ("Month:", calendar.month_name[month]),
         ("Semester:", f"Semester {semester}")],
        [("Instructor:", teacher_name),
         ("Year:", str(year)), ("", "")],
        [("Subject:", subject), ("", ""), ("", "")],
    ]
    for ri, row_data in enumerate(rows_meta):
        r = ri + 2
        ws.row_dimensions[r].height = 15
        col = 1
        for label, val in row_data:
            end_col = min(col + 7, last_col)
            mg(ws, r, col, r, end_col)
            cell = ws.cell(r, col,
                           f"{label}  {val}" if (label or val) else "")
            cell.font = mf(9); cell.alignment = lft
            col = end_col + 1

    # Holiday legend
    leg = max(last_col - 8, NAME_COL + 1)
    mg(ws, 2, leg, 2, last_col)
    lc = ws.cell(2, leg, "  Holiday")
    lc.font = mf(9); lc.fill = fl(SUN_BG); lc.alignment = lft
    mg(ws, 3, leg, 3, last_col)
    lc2 = ws.cell(3, leg, "P- Present;  A- Absent")
    lc2.font = mf(8, color='595959'); lc2.alignment = lft

    ws.row_dimensions[5].height = 5

    # Header rows 6-7
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 16
    mg(ws, 6, NO_COL,   7, NO_COL)
    mg(ws, 6, NAME_COL, 7, NAME_COL)

    def hc(r, col, val, is_sun=False):
        cell = ws.cell(r, col, val)
        cell.font      = mf(9, True, '1F497D' if is_sun else 'FFFFFF')
        cell.fill      = fl(SUN_BG if is_sun else HEADER_BG)
        cell.alignment = ctr; cell.border = bd

    hc(6, NO_COL,   "Roll No")
    hc(6, NAME_COL, "Student Name")
    for d in range(1, num_days+1):
        col = DS + d - 1
        s = d in sundays
        hc(6, col, day_abbrs[d-1], s)
        hc(7, col, d, s)

    # Data rows
    DATA_START = 8
    for idx, (roll, name) in enumerate(students):
        row = DATA_START + idx
        bg  = ALT_BG if idx % 2 == 1 else "FFFFFF"
        ws.row_dimensions[row].height = 15

        rc = ws.cell(row, NO_COL, roll)
        rc.font = mf(8); rc.fill = fl(bg); rc.alignment = ctr; rc.border = bd

        nc = ws.cell(row, NAME_COL, name)
        nc.font = mf(9); nc.fill = fl(bg); nc.alignment = lft; nc.border = bd

        att = attendance_data.get(roll, {})
        for d in range(1, num_days+1):
            col = DS + d - 1
            s   = d in sundays
            val = att.get(d, '')
            ac  = ws.cell(row, col, val)
            ac.font = mf(8, bold=(val == 'P'),
                         color='1F6B2A' if val == 'P'
                               else 'C0392B' if val == 'A' else '000000')
            ac.fill = fl(SUN_BG if s else bg)
            ac.alignment = ctr; ac.border = bd

    ws.column_dimensions[get_column_letter(NO_COL)].width   = 15
    ws.column_dimensions[get_column_letter(NAME_COL)].width = 22
    for d in range(num_days):
        ws.column_dimensions[get_column_letter(DS+d)].width = 4.0

    ws.freeze_panes = ws.cell(DATA_START, DS)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    wb.save(output_path)


# ── Report Window ─────────────────────────────────────────────────────────────

class ReportWindow:
    def __init__(self, parent, container=None):
        W, H = 1280, 820
        if container is not None:
            self.window = container
            self.standalone = False
        else:
            self.standalone = True
            self.window = tk.Toplevel(parent)
            self.window.title("Reports & History")
            _set_app_icon(self.window)
            self.window.resizable(True, True)
            self.window.transient(parent)

            sw = self.window.winfo_screenwidth()
            sh = self.window.winfo_screenheight()
            W = min(1280, sw - 40)
            H = min(820,  sh - 40)
            self.window.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")

        self.W, self.H = W, H

        self.db       = Database()
        self._refs    = {}
        self._records = []   # current displayed records

        self._build_ui()
        self._load_semester_list()
        if self.standalone: self.window.protocol("WM_DELETE_WINDOW", self.window.destroy)

    # ── Background ────────────────────────────────────────────────────────────

    def _load_bg(self, w, h):
        for name in ["tb.jpg", "tb.JPG"]:
            path = os.path.join(get_base_dir(), name)
            if os.path.exists(path):
                img = Image.open(path).resize((w, h), Image.Resampling.LANCZOS)
                img = ImageEnhance.Brightness(img).enhance(0.22)
                img = img.filter(ImageFilter.GaussianBlur(5))
                photo = ImageTk.PhotoImage(img)
                self._refs['bg'] = photo
                return photo
        return None

    def _draw_bg(self):
        w = self.window.winfo_width()  or self.W
        h = self.window.winfo_height() or self.H
        photo = self._load_bg(w, h)
        if photo:
            self.canvas.delete('bg')
            self.canvas.create_image(0, 0, image=photo, anchor='nw', tags='bg')
            self.canvas.tag_lower('bg')

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        self.canvas = tk.Canvas(self.window, highlightthickness=0, bd=0)
        self.canvas.place(x=0, y=0, relwidth=1, relheight=1)
        self.window.bind('<Configure>',
            lambda e: self._draw_bg() if e.widget is self.window else None)
        self._draw_bg()

        # Header
        hdr = GradientFrame(self.window, TM.get('header_grad1','#0d0d2e'), TM.get('header_grad2','#001a1a'), height=62)
        hdr.place(x=0, y=0, relwidth=1)
        def draw_hdr(e, h=hdr):
            h._on_resize(e); h.delete('widgets')
            h.create_text(e.width//2, 31,
                text="📊   ATTENDANCE REPORTS & HISTORY",
                font=("Inter 18pt", 17, "bold"),
                fill='#00E5CC', tags='widgets')
        hdr.bind('<Configure>', draw_hdr)
        tk.Frame(self.window, bg='#00E5CC', height=2).place(x=0, y=62, relwidth=1)

        # ── Filter bar ────────────────────────────────────────────────────────
        fbar = tk.Frame(self.window, bg=TM.get('panel','#080814'), height=58)
        fbar.place(x=0, y=64, relwidth=1)

        lbl_style = dict(font=("Inter 18pt", 9, "bold"),
                         bg=TM.get('panel','#080814'), fg=TM.get('accent2','#7ecdc4'))
        dd_style   = dict(font=("Roboto", 10), bg=TM.get('entry_bg','#1a1a2e'), fg='white',
                          activebackground='#2a2a4e', activeforeground='#00E5CC',
                          relief=tk.FLAT, highlightthickness=0)

        # Semester
        tk.Label(fbar, text="Semester:", **lbl_style).place(x=14, y=14)
        self.sem_var = tk.StringVar(value="All")
        self.sem_menu = tk.OptionMenu(fbar, self.sem_var, "All")
        self.sem_menu.config(**dd_style, width=6)
        self.sem_menu["menu"].config(bg=TM.get('entry_bg','#1a1a2e'), fg='white',
                                      activebackground='#2a2a4e',
                                      font=("Roboto", 9))
        self.sem_menu.place(x=90, y=10)

        # Month
        tk.Label(fbar, text="Month:", **lbl_style).place(x=200, y=14)
        self.month_var = tk.StringVar(value=calendar.month_name[date.today().month])
        months = list(calendar.month_name)[1:]
        self.month_menu = tk.OptionMenu(fbar, self.month_var, *months)
        self.month_menu.config(**dd_style, width=12)
        self.month_menu["menu"].config(bg=TM.get('entry_bg','#1a1a2e'), fg='white',
                                        activebackground='#2a2a4e',
                                        font=("Roboto", 9))
        self.month_menu.place(x=262, y=10)

        # Year
        tk.Label(fbar, text="Year:", **lbl_style).place(x=390, y=14)
        self.year_var = tk.StringVar(value=str(date.today().year))
        cur_y = date.today().year
        years = [str(y) for y in range(cur_y - 1, cur_y + 2)]
        self.year_menu = tk.OptionMenu(fbar, self.year_var, *years)
        self.year_menu.config(**dd_style, width=8)
        self.year_menu["menu"].config(bg=TM.get('entry_bg','#1a1a2e'), fg='white',
                                       activebackground='#2a2a4e',
                                       font=("Roboto", 9))
        self.year_menu.place(x=440, y=10)

        # Load button
        # Load button placed AFTER teacher dropdown
        self.load_btn = GradientButton(fbar,
            text="🔍  Load Records",
            color1='#1a3a6b', color2='#0d2a5a',
            hover_color1='#2980B9', hover_color2='#1a6a9a',
            font=("Inter 18pt", 10, "bold"),
            width=140, height=36,
            command=self._load_records)
        # Teacher filter (placed before Load button)
        tk.Label(fbar, text="Teacher:", **lbl_style).place(x=565, y=14)
        self.teacher_var  = tk.StringVar(value="All")
        self.subject_var  = tk.StringVar(value="All")
        self.teacher_menu2 = tk.OptionMenu(fbar, self.teacher_var, "All")
        self.teacher_menu2.config(**dd_style, width=16)
        self.teacher_menu2["menu"].config(bg=TM.get('entry_bg','#1a1a2e'), fg='white',
            activebackground='#2a2a4e', font=("Roboto",9))
        self.teacher_menu2.place(x=630, y=10)
        # Now place Load button after teacher
        self.load_btn.place(x=808, y=11)

        # Export button
        self.export_btn = GradientButton(fbar,
            text="📥  Export Excel",
            color1='#1a6b3a', color2='#0d4a2a',
            hover_color1='#27AE60', hover_color2='#1a7a4a',
            font=("Inter 18pt", 10, "bold"),
            width=140, height=36,
            command=self._export_excel)
        self.export_btn.place(x=990, y=11)

        tk.Frame(self.window, bg=TM.get('entry_border','#1a1a3e'), height=1).place(x=0, y=122, relwidth=1)

        # ── Teacher info panel ────────────────────────────────────────────────
        info_panel = tk.Frame(self.window, bg=TM.get('panel','#0a0a18'), height=60)
        info_panel.place(x=0, y=123, relwidth=1)
        tk.Frame(self.window, bg=TM.get('entry_border','#1a1a3e'), height=1).place(x=0, y=183, relwidth=1)

        ti_lbl = dict(font=("Inter 18pt", 9, "bold"),
                      bg=TM.get('panel','#0a0a18'), fg=TM.get('accent','#00E5CC'))
        ti_ent = dict(font=("Roboto", 10), bg=TM.get('entry_bg','#1a1a2e'), fg='white',
                      insertbackground='white', relief=tk.FLAT,
                      highlightthickness=1,
                      highlightbackground='#2a2a4e',
                      highlightcolor='#00E5CC')

        tk.Label(info_panel, text="Teacher Name:", **ti_lbl).place(x=14, y=8)
        self.e_teacher = tk.Entry(info_panel, **ti_ent, width=22)
        self.e_teacher.place(x=14, y=28, height=26)

        tk.Label(info_panel, text="Room No:", **ti_lbl).place(x=250, y=8)
        self.e_room = tk.Entry(info_panel, **ti_ent, width=14)
        self.e_room.place(x=250, y=28, height=26)

        tk.Label(info_panel, text="Subject:", **ti_lbl).place(x=420, y=8)
        self.e_subject = tk.Entry(info_panel, **ti_ent, width=22)
        self.e_subject.place(x=420, y=28, height=26)

        tk.Label(info_panel,
            text="ℹ  Fill teacher info before exporting Excel",
            font=("Roboto", 8), bg=TM.get('panel','#0a0a18'), fg='#555577').place(x=660, y=22)

        # ── Stats bar ─────────────────────────────────────────────────────────
        self.stats_frame = GradientFrame(self.window, TM.get('header_grad1','#0d0d2e'), TM.get('header_grad2','#001a1a'), height=36)
        self.stats_frame.place(x=0, y=184, relwidth=1)
        def draw_stats(e, sf=self.stats_frame):
            sf._on_resize(e); sf.delete('widgets')
            sf.create_text(e.width//2, 18,
                text=self._stats_text,
                font=("Roboto", 9), fill='#7ecdc4', tags='widgets')
        self.stats_frame.bind('<Configure>', draw_stats)
        self._stats_text = "Select semester, month & year then click Load Records"
        self._draw_stats = draw_stats

        # ── Table ─────────────────────────────────────────────────────────────
        table_frame = tk.Frame(self.window, bg='#0a0a14')
        table_frame.place(x=0, y=220, relwidth=1, relheight=1, height=-220)

        # Build day columns dynamically (max 31)
        self._day_cols = [str(d) for d in range(1, 32)]
        cols = ('roll', 'name') + tuple(self._day_cols) + ('present', 'absent', 'pct')

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Dark.Treeview',
            background='#0d0d1a', foreground='white',
            fieldbackground='#0d0d1a',
            rowheight=22, font=('Roboto', 9))
        style.configure('Dark.Treeview.Heading',
            background=TM.get('header_grad1','#0d0d2e'), foreground='#00E5CC',
            font=('Inter 18pt', 9, 'bold'), relief='flat')
        style.map('Dark.Treeview',
            background=[('selected', '#1a3a6b')],
            foreground=[('selected', 'white')])

        sb_y = tk.Scrollbar(table_frame, orient=tk.VERTICAL,
                             bg=TM.get('entry_bg','#1a1a2e'), troughcolor='#0d0d1a')
        sb_y.pack(side=tk.RIGHT, fill=tk.Y)
        sb_x = tk.Scrollbar(table_frame, orient=tk.HORIZONTAL,
                             bg=TM.get('entry_bg','#1a1a2e'), troughcolor='#0d0d1a')
        sb_x.pack(side=tk.BOTTOM, fill=tk.X)

        self.tree = ttk.Treeview(table_frame,
            columns=cols, show='headings',
            yscrollcommand=sb_y.set,
            xscrollcommand=sb_x.set,
            style='Dark.Treeview')
        self.tree.pack(fill=tk.BOTH, expand=True)
        sb_y.config(command=self.tree.yview)
        sb_x.config(command=self.tree.xview)

        # Fixed columns
        self.tree.heading('roll', text='Roll Number')
        self.tree.heading('name', text='Student Name')
        self.tree.column('roll', width=120, minwidth=100, anchor='center')
        self.tree.column('name', width=180, minwidth=140, anchor='w')

        # Day columns set up in _load_records based on month
        self.tree.heading('present', text='Present')
        self.tree.heading('absent',  text='Absent')
        self.tree.heading('pct',     text='%')
        self.tree.column('present', width=60, anchor='center')
        self.tree.column('absent',  width=60, anchor='center')
        self.tree.column('pct',     width=55, anchor='center')

        # Alternating row tags
        self.tree.tag_configure('odd',  background='#0d0d1a', foreground='white')
        self.tree.tag_configure('even', background='#111128', foreground='white')
        self.tree.tag_configure('sun',  background='#101828', foreground='#7ecdc4')
        self.tree.tag_configure('present_row',
            background='#0d2a0d', foreground='#2ecc71')

    # ── Data loading ──────────────────────────────────────────────────────────

    def _load_semester_list(self):
        try:
            students = self.db.get_all_students()
            sems = sorted(set(str(s[5]) for s in students))
            menu = self.sem_menu["menu"]
            menu.delete(0, tk.END)
            menu.add_command(label="All",
                command=lambda: self.sem_var.set("All"))
            for s in sems:
                menu.add_command(label=f"Semester {s}",
                    command=lambda v=s: self.sem_var.set(v))
            if sems:
                self.sem_var.set(sems[0])
        except: pass
        try:
            teachers = self.db.get_all_teachers()
            menu2 = self.teacher_menu2["menu"]
            menu2.delete(0, tk.END)
            menu2.add_command(label="All",
                command=lambda: self.teacher_var.set("All"))
            for t in teachers:
                menu2.add_command(label=t,
                    command=lambda v=t: self.teacher_var.set(v))
        except: pass
        # populate subject dropdown with all subjects initially
        self._populate_subjects()

    def _on_teacher_filter_change(self, *_):
        self._populate_subjects()

    def _populate_subjects(self):
        teacher = self.teacher_var.get()
        try:
            menu = self.subject_menu["menu"]
            menu.delete(0, tk.END)
            menu.add_command(label="All",
                command=lambda: self.subject_var.set("All"))
            if teacher == "All":
                self.cursor2 = self.db.conn.cursor()
                self.cursor2.execute(
                    "SELECT DISTINCT subject_name FROM classes ORDER BY subject_name")
                subjects = [r[0] for r in self.cursor2.fetchall()]
            else:
                subjects = [r[0] for r in
                            self.db.get_subjects_for_teacher(teacher)]
            for s in subjects:
                menu.add_command(label=s,
                    command=lambda v=s: self.subject_var.set(v))
            self.subject_var.set("All")
        except Exception as e:
            pass

    def _load_records(self):
        sem_val  = self.sem_var.get()
        month_nm = self.month_var.get()
        year_s   = self.year_var.get()

        try:
            month = list(calendar.month_name).index(month_nm)
            year  = int(year_s)
            sem   = None if sem_val == "All" else int(sem_val)
        except:
            messagebox.showerror("Error", "Invalid month/year selection"); return

        num_days = calendar.monthrange(year, month)[1]
        sundays  = {d for d in range(1, num_days+1)
                    if date(year, month, d).weekday() == 6}

        # Update day column headers
        for d in range(1, 32):
            col = str(d)
            if d <= num_days:
                day_abbr = calendar.day_abbr[
                    date(year, month, d).weekday()][:2]
                is_sun = d in sundays
                self.tree.heading(col,
                    text=f"{day_abbr}\n{d}" if not is_sun else f"☀\n{d}")
                self.tree.column(col, width=28, minwidth=22, anchor='center')
            else:
                self.tree.heading(col, text="")
                self.tree.column(col, width=0, minwidth=0)

        # Get students for semester
        if sem:
            students = self.db.get_students_by_semester(sem)
        else:
            students = self.db.get_all_students()

        if not students:
            messagebox.showinfo("No Data",
                f"No students found for Semester {sem_val}"); return

        # Get attendance for the month
        date_prefix = f"{year}-{month:02d}"
        att_raw = self.db.get_all_attendance()
        # Filter to this month
        month_att = [r for r in att_raw if r[4].startswith(date_prefix)]
        if sem:
            month_att = [r for r in month_att if r[3] == sem]

        # Build dict: roll -> {day: 'P'/'A'}
        att_map = {}
        for rec in month_att:
            # rec: (id, name, roll, semester, date, time, status)
            roll = rec[2]
            try:
                day = int(rec[4].split('-')[2])
            except: continue
            val = 'P' if rec[6] == 'Present' else 'A'
            if roll not in att_map:
                att_map[roll] = {}
            att_map[roll][day] = val

        # Populate tree
        for item in self.tree.get_children():
            self.tree.delete(item)

        self._records = []
        total_p = total_a = 0

        for idx, s in enumerate(students):
            sid, name, _, roll, _, semester_n = s[0],s[1],s[2],s[3],s[4],s[5]
            att = att_map.get(roll, {})

            day_vals = []
            for d in range(1, 32):
                if d <= num_days:
                    day_vals.append(att.get(d, ''))
                else:
                    day_vals.append('')

            present = sum(1 for v in day_vals if v == 'P')
            absent  = sum(1 for v in day_vals if v == 'A')
            total   = present + absent
            pct     = f"{present/total*100:.0f}%" if total > 0 else "-"
            total_p += present
            total_a += absent

            row_data = (roll, name) + tuple(day_vals) + (present, absent, pct)
            tag = 'even' if idx % 2 == 0 else 'odd'
            self.tree.insert('', tk.END, values=row_data, tags=(tag,))
            self._records.append({
                'roll': roll, 'name': name,
                'att': att, 'present': present, 'absent': absent
            })

        # Stats
        self._stats_text = (
            f"Semester {sem_val}  |  "
            f"{calendar.month_name[month]} {year}  |  "
            f"Students: {len(students)}  |  "
            f"Total Present: {total_p}  |  "
            f"Total Absent: {total_a}"
        )
        w = self.stats_frame.winfo_width()
        h = self.stats_frame.winfo_height()
        if w > 1:
            e = type('E', (), {'width': w, 'height': h})()
            self._draw_stats(e)

    # ── Export Excel ──────────────────────────────────────────────────────────

    def _export_excel(self):
        if not self._records:
            messagebox.showwarning("No Data",
                "Load records first before exporting."); return

        teacher = self.e_teacher.get().strip() or "—"
        room    = self.e_room.get().strip()    or "—"
        subject = self.e_subject.get().strip() or "—"

        sem_val  = self.sem_var.get()
        month_nm = self.month_var.get()
        year_s   = self.year_var.get()

        try:
            month = list(calendar.month_name).index(month_nm)
            year  = int(year_s)
            sem   = sem_val
        except:
            messagebox.showerror("Error", "Invalid month/year"); return

        # Default filename
        default_name = (f"Attendance_Sem{sem}_{calendar.month_name[month]}"
                        f"_{year}.xlsx")

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_name,
            title="Save Attendance Report")
        if not filepath:
            return

        # Build student list and attendance map
        students_list = [(r['roll'], r['name']) for r in self._records]
        att_data      = {r['roll']: r['att']   for r in self._records}

        try:
            generate_excel_report(
                output_path   = filepath,
                semester      = sem,
                month         = month,
                year          = year,
                room          = room,
                teacher_name  = teacher,
                subject       = subject,
                students      = students_list,
                attendance_data = att_data
            )
            messagebox.showinfo("Exported",
                f"✅  Report saved!\n{filepath}")
            # Auto-open
            try:
                import subprocess
                subprocess.Popen(f'start "" "{filepath}"', shell=True)
            except: pass
        except Exception as e:
            messagebox.showerror("Export Error", str(e))